import io
from datetime import date, datetime, timedelta
from typing import Dict, List, Any
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.member import Member
from app.models.membership import Membership
from app.models.payment import Payment
from app.models.attendance import AttendanceLog
from app.core.constants import AttendanceSource
from app.models.profile import Profile
from app.models.trainer import Trainer

# Styles configuration
HEADER_FILL = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10, bold=False, color="111111")
SUMMARY_FONT = Font(name="Arial", size=10, bold=True, color="111111")

BORDER_THIN = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0")
)
BORDER_SUMMARY = Border(
    top=Side(style="thin", color="111111"),
    bottom=Side(style="double", color="111111")
)

def _style_worksheet(ws):
    """Apply styling: Freeze top row, auto-fit columns, add borders & font."""
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Auto-fit columns with safety margin
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Calculate max width
        for cell in col:
            # Skip merged or summary cells logic to prevent huge widths
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                
        # Set width
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

class ReportService:
    @staticmethod
    def generate_members_report(db: Session, status_filter: str = "all") -> bytes:
        """
        Excel report: All members with membership status and payment summaries.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Members Catalog"
        
        # Headers
        headers = [
            "Member Name", "Email", "Phone", "Plan", "Start Date", "End Date", 
            "Status", "Days Remaining", "Trainer", "Joining Date", "Total Paid (₹)"
        ]
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER_THIN

        # Query members
        query = db.query(Member).options(
            joinedload(Member.profile).joinedload(Profile.user),
            joinedload(Member.memberships).joinedload(Membership.plan)
        ).filter(Member.is_deleted == False)
        
        members = query.all()
        row_idx = 2
        
        for m in members:
            # Filter by status if requested
            active_m = m.active_membership
            m_status = active_m.status if active_m else "No Plan"
            
            if status_filter != "all" and status_filter.lower() != m_status.lower():
                continue

            # Calculate total paid
            total_paid = db.query(func.sum(Payment.amount)).filter(
                Payment.member_id == m.id,
                Payment.status == "success"
            ).scalar() or 0.0

            # Get assigned trainer name
            trainer_name = "N/A"
            if hasattr(m, "assigned_trainers") and m.assigned_trainers:
                active_trainer = [at for at in m.assigned_trainers if at.is_active]
                if active_trainer:
                    trainer_name = active_trainer[0].trainer.profile.full_name

            # Write row
            row_data = [
                m.profile.full_name,
                m.profile.user.email,
                m.profile.phone or "N/A",
                active_m.plan_name if active_m else "N/A",
                active_m.start_date.strftime("%d-%b-%Y") if active_m else "N/A",
                active_m.end_date.strftime("%d-%b-%Y") if active_m else "N/A",
                m_status.upper(),
                active_m.days_remaining if active_m else 0,
                trainer_name,
                m.joining_date.strftime("%d-%b-%Y") if m.joining_date else "N/A",
                float(total_paid)
            ]

            fill_to_use = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.fill = fill_to_use
                cell.border = BORDER_THIN
                
                # Alignments and number formats
                if col_idx in [5, 6, 10]: # Dates
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [7, 8]: # Status / Days remaining
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx == 11: # Price Currency
                    cell.number_format = "₹#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                    
            row_idx += 1

        _style_worksheet(ws)
        
        # Save to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    @staticmethod
    def generate_attendance_report(db: Session, from_date: date, to_date: date) -> bytes:
        """
        Excel report: Attendance log for date range.
        Sheet 1 - Daily Summary: Date, Total Visits, Unique Members, Biometric Count, Manual Count
        Sheet 2 - Member Detail: Member Name, Date, Check-in, Check-out, Duration, Source
        """
        wb = Workbook()
        
        # Sheet 1: Daily Summary
        ws1 = wb.active
        ws1.title = "Daily Summary"
        
        s_headers = ["Date", "Total Visits", "Unique Members Checked In", "Biometric Scans", "Manual Logs"]
        for col_idx, text in enumerate(s_headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_THIN

        # Sheet 2: Member Detail
        ws2 = wb.create_sheet(title="Attendance Logs Detail")
        d_headers = ["Member Name", "Date", "Check-in Time", "Check-out Time", "Duration (Mins)", "Source"]
        for col_idx, text in enumerate(d_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_THIN

        # Query attendance records
        records = db.query(AttendanceLog).options(
            joinedload(AttendanceLog.member).joinedload(Member.profile)
        ).filter(
            AttendanceLog.check_in >= datetime.combine(from_date, datetime.min.time()),
            AttendanceLog.check_in <= datetime.combine(to_date, datetime.max.time())
        ).order_by(AttendanceLog.check_in.desc()).all()

        # Build Sheet 1 Summaries
        summary_map = {}
        curr_d = from_date
        while curr_d <= to_date:
            summary_map[curr_d] = {"visits": 0, "members": set(), "biometric": 0, "manual": 0}
            curr_d += timedelta(days=1)

        row_idx_2 = 2
        for r in records:
            r_date = r.check_in.date()
            if r_date in summary_map:
                summary_map[r_date]["visits"] += 1
                if r.member_id:
                    summary_map[r_date]["members"].add(r.member_id)
                # Biometric vs Manual source
                is_bio = r.source == AttendanceSource.BIOMETRIC
                if is_bio:
                    summary_map[r_date]["biometric"] += 1
                else:
                    summary_map[r_date]["manual"] += 1

            # Populate Detail Sheet (Sheet 2)
            duration_mins = "N/A"
            if r.check_out:
                duration_mins = int((r.check_out - r.check_in).total_seconds() / 60)

            member_name = r.member.profile.full_name if r.member else "Unknown"
            
            row_data_2 = [
                member_name,
                r_date.strftime("%d-%b-%Y"),
                r.check_in.strftime("%I:%M %p"),
                r.check_out.strftime("%I:%M %p") if r.check_out else "N/A",
                duration_mins,
                "BIOMETRIC" if r.source == AttendanceSource.BIOMETRIC else "MANUAL"
            ]

            fill_to_use = ZEBRA_FILL if row_idx_2 % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate(row_data_2, 1):
                cell = ws2.cell(row=row_idx_2, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.fill = fill_to_use
                cell.border = BORDER_THIN
                if col_idx in [2, 3, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center")
            row_idx_2 += 1

        # Populate Summary Sheet (Sheet 1)
        row_idx_1 = 2
        for s_date, s in sorted(summary_map.items()):
            row_data_1 = [
                s_date.strftime("%d-%b-%Y"),
                s["visits"],
                len(s["members"]),
                s["biometric"],
                s["manual"]
            ]
            fill_to_use = ZEBRA_FILL if row_idx_1 % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate(row_data_1, 1):
                cell = ws1.cell(row=row_idx_1, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.fill = fill_to_use
                cell.border = BORDER_THIN
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
            row_idx_1 += 1

        _style_worksheet(ws1)
        _style_worksheet(ws2)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    @staticmethod
    def generate_payments_report(db: Session, from_date: date, to_date: date) -> bytes:
        """
        Excel report: All payment transactions with receipt numbers and collection details.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments & Transactions"

        headers = [
            "Receipt No", "Member Name", "Plan Purchased", "Amount (₹)", 
            "Payment Method", "Status", "Date", "Collected By", "Transaction Ref"
        ]

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_THIN

        payments = db.query(Payment).options(
            joinedload(Payment.member).joinedload(Member.profile),
            joinedload(Payment.membership).joinedload(Membership.plan)
        ).filter(
            func.date(Payment.created_at) >= from_date,
            func.date(Payment.created_at) <= to_date,
            Payment.is_deleted == False
        ).order_by(Payment.created_at.desc()).all()

        row_idx = 2
        total_collected = 0.0
        method_summary = {}

        for p in payments:
            amount = float(p.amount)
            if p.status == "success":
                total_collected += amount
                method = p.payment_method or "other"
                method_summary[method] = method_summary.get(method, 0.0) + amount

            member_name = p.member.profile.full_name if p.member else "N/A"
            plan_name = p.membership.plan.name if (p.membership and p.membership.plan) else "N/A"
            
            row_data = [
                p.receipt_number or "N/A",
                member_name,
                plan_name,
                amount,
                p.payment_method.upper() if p.payment_method else "N/A",
                p.status.upper(),
                p.created_at.strftime("%d-%b-%Y"),
                "SYSTEM",  # or collected_by
                p.transaction_reference or "N/A"
            ]

            fill_to_use = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.fill = fill_to_use
                cell.border = BORDER_THIN
                
                if col_idx == 4: # Amount
                    cell.number_format = "₹#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center")

            row_idx += 1

        # Summary rows at bottom
        row_idx += 1
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        cell = ws.cell(row=row_idx, column=1, value="TOTAL COLLECTED (SUCCESSFUL)")
        cell.font = SUMMARY_FONT
        cell.alignment = Alignment(horizontal="right")
        
        sum_cell = ws.cell(row=row_idx, column=4, value=total_collected)
        sum_cell.font = SUMMARY_FONT
        sum_cell.number_format = "₹#,##0.00"
        sum_cell.alignment = Alignment(horizontal="right")
        sum_cell.border = BORDER_SUMMARY

        # Method breakdowns
        row_idx += 1
        for m_name, m_total in sorted(method_summary.items()):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            ws.cell(row=row_idx, column=1, value=f"Total via {m_name.upper()}").font = DATA_FONT
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="right")
            
            m_cell = ws.cell(row=row_idx, column=4, value=m_total)
            m_cell.font = DATA_FONT
            m_cell.number_format = "₹#,##0.00"
            m_cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        _style_worksheet(ws)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    @staticmethod
    def generate_expiry_report(db: Session, days_ahead: int = 30) -> bytes:
        """
        Excel report: Members whose memberships are expiring in the coming days.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Expiring Memberships"

        headers = [
            "Member Name", "Phone", "Email", "Active Plan", "Expiry Date", "Days Remaining", "Assigned Trainer"
        ]

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER_THIN

        today = date.today()
        target_date = today + timedelta(days=days_ahead)

        expiring_memberships = db.query(Membership).options(
            joinedload(Membership.plan),
            joinedload(Membership.member).joinedload(Member.profile).joinedload(Profile.user)
        ).filter(
            Membership.status == "active",
            Membership.end_date >= today,
            Membership.end_date <= target_date,
            Membership.is_deleted == False
        ).order_by(Membership.end_date.asc()).all()

        row_idx = 2
        for m in expiring_memberships:
            member = m.member
            days_left = (m.end_date - today).days

            trainer_name = "N/A"
            if member and hasattr(member, "assigned_trainers") and member.assigned_trainers:
                active_trainer = [at for at in member.assigned_trainers if at.is_active]
                if active_trainer:
                    trainer_name = active_trainer[0].trainer.profile.full_name

            row_data = [
                member.profile.full_name if member else "Unknown",
                member.profile.phone if (member and member.profile.phone) else "N/A",
                member.profile.user.email if (member and member.profile.user) else "N/A",
                m.plan.name if m.plan else "N/A",
                m.end_date.strftime("%d-%b-%Y"),
                days_left,
                trainer_name
            ]

            fill_to_use = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.fill = fill_to_use
                cell.border = BORDER_THIN
                if col_idx in [5, 6]:
                    cell.alignment = Alignment(horizontal="center")

            row_idx += 1

        _style_worksheet(ws)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()
